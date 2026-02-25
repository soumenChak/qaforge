"""
QAForge -- Template Engine
============================
Renders test cases into various output formats (Excel, JSON, Markdown).

Excel output uses openpyxl with FreshGravity branding:
  - FG teal header (#2BB8C6) with white bold text
  - Auto-column widths
  - Frozen first row
  - Sheet name: "Test Cases"
"""

from __future__ import annotations

import json
import logging
from pathlib import Path
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)

# FreshGravity brand colour
FG_TEAL = "2BB8C6"
FG_WHITE = "FFFFFF"

# ---------------------------------------------------------------------------
# Default column templates per domain
# ---------------------------------------------------------------------------

_DEFAULT_COLUMNS = [
    {"key": "test_case_id",   "header": "TC ID",           "width": 16},
    {"key": "title",          "header": "Title",           "width": 40},
    {"key": "description",    "header": "Description",     "width": 55},
    {"key": "preconditions",  "header": "Preconditions",   "width": 40},
    {"key": "test_steps",     "header": "Steps",           "width": 65},
    {"key": "expected_result", "header": "Expected Result", "width": 45},
    {"key": "priority",       "header": "Priority",        "width": 12},
    {"key": "category",       "header": "Category",        "width": 18},
    {"key": "status",         "header": "Status",          "width": 14},
]

_MDM_EXTRA_COLUMNS = [
    {"key": "domain_tags", "header": "Domain Tags", "width": 30},
]


def get_default_template(domain: str = "generic") -> List[Dict[str, Any]]:
    """
    Return the default column mapping for a given domain.

    Args:
        domain: The domain identifier (e.g. "mdm", "generic").

    Returns:
        A list of column config dicts with keys: key, header, width.
    """
    columns = list(_DEFAULT_COLUMNS)
    if domain.lower() in ("mdm", "reltio", "semarchy"):
        columns.extend(_MDM_EXTRA_COLUMNS)
    return columns


# ---------------------------------------------------------------------------
# Excel renderer
# ---------------------------------------------------------------------------

def render_excel(
    test_cases: List[Dict[str, Any]],
    output_path: str | Path,
    template_config: Optional[List[Dict[str, Any]]] = None,
    sheet_name: str = "Test Cases",
) -> Path:
    """
    Generate an Excel file from test cases with FreshGravity branding.

    Args:
        test_cases: List of test-case dicts.
        output_path: Path to write the .xlsx file.
        template_config: Column definitions (from get_default_template).
            If None, uses the generic default template.
        sheet_name: Excel sheet name (default: "Test Cases").

    Returns:
        The Path to the written Excel file.

    Raises:
        ImportError: If openpyxl is not installed.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required for Excel export. "
            "Run: pip install openpyxl>=3.1.0"
        ) from exc

    output_path = Path(output_path)
    columns = template_config or get_default_template()

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # --- Styles ---
    header_fill = PatternFill(start_color=FG_TEAL, end_color=FG_TEAL, fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color=FG_WHITE)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    # --- Header row ---
    for col_idx, col_def in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_def["header"])
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # --- Data rows ---
    for row_idx, tc in enumerate(test_cases, start=2):
        for col_idx, col_def in enumerate(columns, start=1):
            key = col_def["key"]
            value = tc.get(key, "")

            # Format complex fields
            if key == "test_steps":
                value = _format_steps_for_excel(value)
            elif key == "domain_tags":
                if isinstance(value, list):
                    value = ", ".join(str(v) for v in value)
            elif key == "status":
                # Default status for new test cases
                value = tc.get("status", "Not Executed")

            cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value else "")
            cell.alignment = cell_alignment
            cell.border = thin_border

    # --- Column widths ---
    for col_idx, col_def in enumerate(columns, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = col_def.get("width", 20)

    # --- Freeze first row ---
    ws.freeze_panes = "A2"

    # --- Auto-filter ---
    last_col = get_column_letter(len(columns))
    last_row = len(test_cases) + 1
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"

    # --- Save ---
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    logger.info("Excel file written: %s (%d test cases)", output_path, len(test_cases))
    return output_path


def _format_steps_for_excel(steps: Any) -> str:
    """
    Format test steps for display in an Excel cell.

    Converts a list of step dicts into a numbered, readable string.
    """
    if not steps:
        return ""
    if isinstance(steps, str):
        return steps

    lines: List[str] = []
    for step in steps:
        if isinstance(step, dict):
            num = step.get("step_number", "")
            action = step.get("action", "")
            expected = step.get("expected_result", "")
            line = f"{num}. {action}"
            if expected:
                line += f"\n   Expected: {expected}"
            lines.append(line)
        elif isinstance(step, str):
            lines.append(step)
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# JSON renderer
# ---------------------------------------------------------------------------

def render_json(
    test_cases: List[Dict[str, Any]],
    indent: int = 2,
    include_metadata: bool = False,
    metadata: Optional[Dict[str, Any]] = None,
) -> str:
    """
    Render test cases as a JSON string.

    Args:
        test_cases: List of test-case dicts.
        indent: JSON indentation level.
        include_metadata: If True, wraps in {metadata, test_cases}.
        metadata: Optional metadata dict to include.

    Returns:
        JSON string.
    """
    if include_metadata and metadata:
        output = {
            "metadata": metadata,
            "test_cases": test_cases,
            "count": len(test_cases),
        }
    else:
        output = test_cases

    return json.dumps(output, indent=indent, ensure_ascii=False, default=str)


# ---------------------------------------------------------------------------
# Markdown renderer
# ---------------------------------------------------------------------------

def render_markdown(
    test_cases: List[Dict[str, Any]],
    title: str = "QAForge Test Cases",
) -> str:
    """
    Render test cases as a Markdown document.

    Args:
        test_cases: List of test-case dicts.
        title: Document title.

    Returns:
        Markdown string.
    """
    lines: List[str] = [f"# {title}\n"]
    lines.append(f"**Total:** {len(test_cases)} test cases\n")

    for tc in test_cases:
        tc_id = tc.get("test_case_id", "TC-???")
        title_text = tc.get("title", "Untitled")
        lines.append(f"## {tc_id}: {title_text}")
        lines.append("")
        lines.append(f"| Field | Value |")
        lines.append(f"|-------|-------|")
        lines.append(f"| **Priority** | {tc.get('priority', 'Medium')} |")
        lines.append(f"| **Category** | {tc.get('category', 'Functional')} |")

        tags = tc.get("domain_tags", [])
        if tags:
            lines.append(f"| **Tags** | {', '.join(tags)} |")

        lines.append("")
        lines.append(f"**Description:** {tc.get('description', '')}")
        lines.append("")
        lines.append(f"**Preconditions:** {tc.get('preconditions', 'None')}")
        lines.append("")

        steps = tc.get("test_steps", [])
        if steps:
            lines.append("**Steps:**")
            lines.append("")
            for step in steps:
                if isinstance(step, dict):
                    num = step.get("step_number", "")
                    action = step.get("action", "")
                    expected = step.get("expected_result", "")
                    lines.append(f"{num}. {action}")
                    if expected:
                        lines.append(f"   - *Expected:* {expected}")

        lines.append("")
        lines.append(f"**Expected Result:** {tc.get('expected_result', '')}")
        lines.append("")
        lines.append("---")
        lines.append("")

    return "\n".join(lines)
