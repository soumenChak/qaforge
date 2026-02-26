"""
QAForge -- Test case management routes.

Prefix: /api/projects/{project_id}/test-cases
(Note: main.py maps "test_cases" -> "/api/test-cases", but these routes
include the project_id in their path definitions.)

Endpoints:
    POST   /generate        — trigger AI generation pipeline
    GET    /                 — list test cases (filter, paginate)
    POST   /                 — add manual test case
    GET    /{tc_id}          — single test case detail
    PUT    /{tc_id}          — update test case
    DELETE /{tc_id}          — delete test case
    POST   /{tc_id}/rate     — rate a test case (1-5 + feedback)
    POST   /export           — export test cases to Excel (file download)
"""

import io
import json
import logging
import time
import uuid
from typing import Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, Request, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy import func, or_
from sqlalchemy.orm import Session

from db_models import (
    FeedbackEntry,
    GenerationRun,
    KnowledgeEntry,
    Project,
    Requirement,
    TestCase,
    TestTemplate,
    User,
)
from db_session import get_db
from dependencies import (
    audit_log,
    get_client_ip,
    get_current_user,
    sanitize_string,
    track_cost,
)
from models import (
    ChatGenerateRequest,
    ChatGenerateResponse,
    ChatMessage,
    MessageResponse,
    TestCaseCreate,
    TestCaseExportRequest,
    TestCaseGenerateRequest,
    TestCaseRateRequest,
    TestCaseResponse,
    TestCaseUpdate,
)

logger = logging.getLogger(__name__)

router = APIRouter()


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _get_project_or_404(project_id: uuid.UUID, db: Session) -> Project:
    project = db.query(Project).filter(Project.id == project_id).first()
    if project is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="Project not found")
    return project


def _get_test_case_or_404(
    tc_id: uuid.UUID, project_id: uuid.UUID, db: Session
) -> TestCase:
    tc = db.query(TestCase).filter(
        TestCase.id == tc_id,
        TestCase.project_id == project_id,
    ).first()
    if tc is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="Test case not found")
    return tc


# ---------------------------------------------------------------------------
# POST /{project_id}/test-cases/generate
# ---------------------------------------------------------------------------
@router.post(
    "/{project_id}/test-cases/generate",
    response_model=list[TestCaseResponse],
    summary="Generate test cases using AI",
    status_code=status.HTTP_201_CREATED,
)
def generate_test_cases(
    project_id: uuid.UUID,
    body: TestCaseGenerateRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Trigger the AI test case generation pipeline.

    Accepts a description and optional requirement_ids. Calls the pipeline
    orchestrator to generate test cases. Falls back to mock generation if
    the pipeline is not yet available.
    """
    project = _get_project_or_404(project_id, db)
    start_time = time.time()

    # ── Gather requirement context ──
    # If specific IDs provided, use those; otherwise auto-fetch ALL project requirements
    requirement_context = ""
    if body.requirement_ids:
        reqs = (
            db.query(Requirement)
            .filter(
                Requirement.id.in_(body.requirement_ids),
                Requirement.project_id == project_id,
            )
            .all()
        )
    else:
        # Auto-fetch all active requirements for the project
        reqs = (
            db.query(Requirement)
            .filter(
                Requirement.project_id == project_id,
                Requirement.status != "deferred",
            )
            .order_by(Requirement.created_at.asc())
            .limit(50)
            .all()
        )

    if reqs:
        requirement_context = "\n".join(
            f"- [{r.priority.upper()}] {r.req_id}: {r.title} -- {r.description or ''}" for r in reqs
        )
        logger.info("Requirements context: %d entries for project %s", len(reqs), project.name)

    # ── Gather BRD/PRD document text ──
    brd_prd_context = ""
    if body.brd_prd_text and body.brd_prd_text.strip():
        brd_prd_context = body.brd_prd_text.strip()[:8000]
        logger.info("BRD/PRD context: %d chars", len(brd_prd_context))

    # ── Gather reference test cases ──
    reference_tc_context = ""
    if body.reference_test_case_ids:
        ref_tcs = (
            db.query(TestCase)
            .filter(
                TestCase.id.in_(body.reference_test_case_ids),
                TestCase.project_id == project_id,
            )
            .limit(10)
            .all()
        )
        if ref_tcs:
            ref_lines = []
            for tc in ref_tcs:
                steps_summary = ""
                if tc.test_steps:
                    steps_summary = " | ".join(
                        f"Step {s.get('step_number', '?')}: {s.get('action', '')}"
                        for s in tc.test_steps[:5]
                    )
                ref_lines.append(
                    f"- {tc.test_case_id} [{tc.execution_type}] {tc.title}\n"
                    f"  Description: {tc.description or 'N/A'}\n"
                    f"  Steps: {steps_summary or 'N/A'}\n"
                    f"  Expected: {tc.expected_result or 'N/A'}"
                )
            reference_tc_context = "\n".join(ref_lines)
            logger.info("Reference TCs: %d loaded", len(ref_tcs))

    # Extract app profile for generation context
    app_profile = project.app_profile

    # Try the real pipeline first
    generated = _generate_via_pipeline(
        description=body.description,
        domain=body.domain,
        sub_domain=body.sub_domain,
        requirement_context=requirement_context,
        count=body.count,
        additional_context=body.additional_context,
        project_id=project_id,
        user_id=current_user.id,
        db=db,
        brd_prd_context=brd_prd_context,
        reference_tc_context=reference_tc_context,
        app_profile=app_profile,
    )

    # Save generated test cases to DB
    base_count = db.query(TestCase).filter(
        TestCase.project_id == project_id
    ).count()

    created = []
    for i, item in enumerate(generated):
        tc_id_str = f"TC-{base_count + i + 1:03d}"

        # Avoid duplicate tc_id
        existing = db.query(TestCase).filter(
            TestCase.project_id == project_id,
            TestCase.test_case_id == tc_id_str,
        ).first()
        if existing:
            tc_id_str = f"TC-{base_count + i + 100:03d}"

        # Normalise fields that the LLM may return as list instead of str
        def _to_str(val):
            if val is None:
                return None
            if isinstance(val, list):
                return "; ".join(str(v) for v in val)
            return str(val)

        # Determine execution_type: from LLM output -> from request body -> default "api"
        exec_type = item.get("execution_type", body.execution_type or "api")
        if exec_type not in ("api", "ui", "sql", "manual"):
            exec_type = "api"

        tc = TestCase(
            project_id=project_id,
            test_case_id=tc_id_str,
            title=sanitize_string(_to_str(item.get("title"))) or "Generated Test Case",
            description=sanitize_string(_to_str(item.get("description"))),
            preconditions=sanitize_string(_to_str(item.get("preconditions"))),
            test_steps=item.get("test_steps"),
            expected_result=sanitize_string(_to_str(item.get("expected_result"))),
            test_data=item.get("test_data"),
            priority=item.get("priority", body.priority or "P2"),
            category=item.get("category", body.category or "functional"),
            execution_type=exec_type,
            domain_tags=item.get("domain_tags"),
            source="ai_generated",
            status="draft",
            generated_by_model=item.get("model", "mock"),
            generation_metadata=item.get("metadata"),
            created_by=current_user.id,
        )
        db.add(tc)
        db.flush()
        created.append(tc)

    duration = time.time() - start_time

    # Track generation run
    gen_run = GenerationRun(
        project_id=project_id,
        agent_type=body.domain,
        input_context={
            "description": body.description[:500],
            "domain": body.domain,
            "sub_domain": body.sub_domain,
            "count": body.count,
        },
        test_cases_generated=len(created),
        llm_provider=generated[0].get("provider", "mock") if generated else "mock",
        llm_model=generated[0].get("model", "mock") if generated else "mock",
        duration_seconds=round(duration, 2),
    )
    db.add(gen_run)
    db.flush()

    audit_log(
        db,
        user_id=current_user.id,
        action="generate_test_cases",
        entity_type="test_case",
        entity_id=str(project_id),
        details={
            "count_requested": body.count,
            "count_generated": len(created),
            "duration_seconds": round(duration, 2),
        },
        ip_address=get_client_ip(request),
    )

    logger.info(
        "Generated %d test cases for project %s in %.1fs",
        len(created),
        project.name,
        duration,
    )

    return [TestCaseResponse.model_validate(tc) for tc in created]


def _generate_via_pipeline(
    description: str,
    domain: str,
    sub_domain: str,
    requirement_context: str,
    count: int,
    additional_context: Optional[str],
    project_id: uuid.UUID,
    user_id: uuid.UUID,
    db: Session,
    brd_prd_context: str = "",
    reference_tc_context: str = "",
    app_profile: Optional[dict] = None,
) -> list[dict]:
    """
    Call the pipeline orchestrator to generate test cases.

    Falls back to direct LLM generation if the pipeline is not importable.
    Falls back to mock generation if LLM also fails.
    """
    # Try importing the real pipeline (async orchestrator)
    try:
        import asyncio
        from pipeline.orchestrator import GenerateRequest, Orchestrator

        req = GenerateRequest(
            description=description,
            domain=domain,
            sub_domain=sub_domain,
            requirements=[requirement_context] if requirement_context else None,
            count=count,
            additional_context=additional_context or "",
        )
        loop = asyncio.new_event_loop()
        try:
            gen_result = loop.run_until_complete(Orchestrator().run(req))
        finally:
            loop.close()

        if gen_result and gen_result.test_cases:
            return gen_result.test_cases
    except ImportError:
        logger.info("pipeline.orchestrator not available; using direct LLM generation")
    except Exception:
        logger.warning("Pipeline generation failed; trying direct LLM", exc_info=True)

    # Try LLM-based generation directly
    try:
        return _generate_via_llm(
            description=description,
            domain=domain,
            sub_domain=sub_domain,
            requirement_context=requirement_context,
            count=count,
            additional_context=additional_context,
            project_id=project_id,
            user_id=user_id,
            db=db,
            brd_prd_context=brd_prd_context,
            reference_tc_context=reference_tc_context,
            app_profile=app_profile,
        )
    except Exception:
        logger.warning("LLM generation failed; using mock fallback", exc_info=True)

    # Mock fallback
    return _mock_generate(description, domain, sub_domain, count)


def _generate_via_llm(
    description: str,
    domain: str,
    sub_domain: str,
    requirement_context: str,
    count: int,
    additional_context: Optional[str],
    project_id: uuid.UUID,
    user_id: uuid.UUID,
    db: Session,
    brd_prd_context: str = "",
    reference_tc_context: str = "",
    app_profile: Optional[dict] = None,
) -> list[dict]:
    """Generate test cases by calling the LLM provider directly (uses smart model)."""
    from core.llm_provider import get_llm_provider

    # ── Query Knowledge Base for domain-relevant context ──
    kb_entries = (
        db.query(KnowledgeEntry)
        .filter(
            or_(
                KnowledgeEntry.domain == domain,
                KnowledgeEntry.domain == "general",
            )
        )
        .order_by(KnowledgeEntry.usage_count.desc(), KnowledgeEntry.created_at.desc())
        .limit(8)
        .all()
    )

    kb_context = ""
    kb_count = 0
    if kb_entries:
        kb_lines = []
        for entry in kb_entries:
            kb_lines.append(
                f"[{entry.entry_type.upper()}] {entry.title}\n{entry.content[:500]}"
            )
            entry.usage_count += 1
        kb_context = "\n\n".join(kb_lines)
        kb_count = len(kb_entries)
        db.flush()
        logger.info("KB context: %d entries for domain=%s", kb_count, domain)

    system_prompt = f"""You are a senior QA automation engineer specializing in {domain} / {sub_domain}.
You generate EXECUTION-READY test cases — each test case must contain enough detail in its test_steps
that an automated execution engine can run it without human intervention.

You understand three execution modes:
1. **API Testing** — HTTP requests with specific endpoints, methods, status codes, expected response fields
2. **UI Testing (Playwright)** — Browser automation with CSS selectors, navigation URLs, form fills, assertions
3. **SQL Testing** — Database queries with specific SQL, expected row counts, column checks

Your test steps must be CONCRETE and SPECIFIC — never vague like "verify the page works".
Instead: "GET /api/users returns 200 with fields: id, name, email" or "Click button#submit, assert .success-toast is visible".

IMPORTANT: You have been provided with the system description, requirements from BRD/PRD documents,
and possibly reference test cases. Use ALL of this context to generate high-quality, domain-specific
test cases that thoroughly cover the described system's functionality, edge cases, and integration points.

Return ONLY a JSON array — no markdown fences, no explanation."""

    # ── Build rich user prompt with all available context ──
    user_prompt_parts = [f"Generate exactly {count} EXECUTION-READY test cases for the following system."]

    user_prompt_parts.append(f"\n=== SYSTEM DESCRIPTION ===\n{description[:4000]}")

    # ── Inject Application Profile (CRITICAL for generation accuracy) ──
    if app_profile and isinstance(app_profile, dict):
        ap_parts = []

        if app_profile.get("app_url"):
            ap_parts.append(f"Application URL: {app_profile['app_url']}")
        if app_profile.get("api_base_url"):
            ap_parts.append(f"API Base URL: {app_profile['api_base_url']}")

        if app_profile.get("tech_stack") and isinstance(app_profile["tech_stack"], dict):
            ts = app_profile["tech_stack"]
            stack_str = ", ".join(f"{k}: {v}" for k, v in ts.items() if v)
            if stack_str:
                ap_parts.append(f"Tech Stack: {stack_str}")

        if app_profile.get("auth") and isinstance(app_profile["auth"], dict):
            auth = app_profile["auth"]
            auth_lines = []
            if auth.get("login_endpoint"):
                auth_lines.append(f"Login: {auth['login_endpoint']}")
            if auth.get("request_body"):
                auth_lines.append(f"Body format: {auth['request_body']}")
            if auth.get("token_header"):
                auth_lines.append(f"Auth header: {auth['token_header']}")
            if auth.get("test_credentials") and isinstance(auth["test_credentials"], dict):
                creds = auth["test_credentials"]
                auth_lines.append(f"Test credentials: {json.dumps(creds)}")
            if auth.get("response_fields"):
                auth_lines.append(f"Login response fields: {', '.join(auth['response_fields'])}")
            if auth_lines:
                ap_parts.append("Authentication:\n  " + "\n  ".join(auth_lines))

        if app_profile.get("api_endpoints") and isinstance(app_profile["api_endpoints"], list):
            ep_lines = []
            for ep in app_profile["api_endpoints"][:30]:
                if not isinstance(ep, dict):
                    continue
                line = f"{ep.get('method', 'GET')} {ep.get('path', '/')}"
                if ep.get("description"):
                    line += f" -- {ep['description']}"
                if ep.get("required_fields") and isinstance(ep["required_fields"], list):
                    line += f" (required: {', '.join(ep['required_fields'])})"
                if ep.get("response_fields") and isinstance(ep["response_fields"], list):
                    line += f" (returns: {', '.join(ep['response_fields'])})"
                ep_lines.append(line)
            if ep_lines:
                ap_parts.append("Known API Endpoints:\n  " + "\n  ".join(ep_lines))

        if app_profile.get("ui_pages") and isinstance(app_profile["ui_pages"], list):
            page_lines = []
            for pg in app_profile["ui_pages"][:20]:
                if not isinstance(pg, dict):
                    continue
                line = f"{pg.get('route', '/')}"
                if pg.get("description"):
                    line += f" -- {pg['description']}"
                if pg.get("key_elements") and isinstance(pg["key_elements"], list):
                    line += f" (selectors: {', '.join(pg['key_elements'])})"
                page_lines.append(line)
            if page_lines:
                ap_parts.append("Known UI Pages:\n  " + "\n  ".join(page_lines))

        if app_profile.get("rbac_model"):
            ap_parts.append(f"RBAC Model: {app_profile['rbac_model']}")

        if app_profile.get("notes"):
            ap_parts.append(f"Important Notes: {app_profile['notes']}")

        if ap_parts:
            app_profile_text = "\n".join(ap_parts)
            user_prompt_parts.append(
                f"\n=== APPLICATION PROFILE (use these EXACT URLs, endpoints, field names, and selectors) ===\n"
                f"CRITICAL: Do NOT invent or guess URLs, endpoints, field names, CSS selectors, or auth flows.\n"
                f"Use ONLY the information below. If an endpoint or page is not listed, do not fabricate one.\n\n"
                f"{app_profile_text[:5000]}"
            )
            logger.info("App profile context injected: %d chars", len(app_profile_text[:5000]))

    if brd_prd_context:
        user_prompt_parts.append(
            f"\n=== BRD/PRD DOCUMENT CONTEXT ===\n"
            f"The following is extracted from the project's BRD/PRD documents. "
            f"Use this to understand the business requirements, acceptance criteria, "
            f"and expected system behavior:\n{brd_prd_context[:6000]}"
        )

    if requirement_context:
        user_prompt_parts.append(
            f"\n=== REQUIREMENTS / USE CASES ===\n"
            f"Generate test cases that cover these specific requirements:\n{requirement_context[:4000]}"
        )

    if reference_tc_context:
        user_prompt_parts.append(
            f"\n=== REFERENCE TEST CASES (match this style and level of detail) ===\n"
            f"Use these existing test cases as examples of the expected format, "
            f"detail level, and naming conventions:\n{reference_tc_context[:3000]}"
        )

    if additional_context:
        user_prompt_parts.append(f"\n=== ADDITIONAL CONTEXT ===\n{additional_context[:2000]}")

    if kb_context:
        user_prompt_parts.append(
            f"\n=== KNOWLEDGE BASE REFERENCE (use these patterns to improve test quality) ===\n{kb_context}"
        )

    user_prompt_parts.append(f"""
For each test case, return a JSON object with ALL of these fields:

- "title": Concise test case title (e.g. "Verify user login API returns JWT token")
- "description": What this test validates and why it matters
- "preconditions": Setup required (e.g. "Valid user exists with email test@example.com")
- "test_steps": Array of step objects — MUST be execution-ready (see format below)
- "expected_result": Overall expected outcome
- "priority": One of "P1", "P2", "P3", "P4"
- "category": One of "functional", "integration", "regression", "smoke", "e2e"
- "execution_type": One of "api", "ui", "sql"
- "domain_tags": Array of relevant tags

=== CRITICAL: test_steps FORMAT by execution_type ===

For "api" test cases, each step MUST reference specific HTTP details:
  {{"step_number": 1, "action": "POST /api/auth/login with body {{\\"email\\":\\"test@example.com\\",\\"password\\":\\"pass123\\"}}", "expected_result": "Returns 200 with fields: access_token, user"}}
  {{"step_number": 2, "action": "GET /api/users with Authorization: Bearer <token>", "expected_result": "Returns 200 with array of user objects containing id, name, email"}}

For "ui" test cases, each step MUST use Playwright-compatible actions and CSS selectors:
  {{"step_number": 1, "action": "Navigate to /login", "expected_result": "Login page loads with email and password fields visible"}}
  {{"step_number": 2, "action": "Fill input#email with 'admin@example.com'", "expected_result": "Email field populated"}}
  {{"step_number": 3, "action": "Fill input[type=password] with 'password123'", "expected_result": "Password field populated"}}
  {{"step_number": 4, "action": "Click button[type=submit]", "expected_result": "Form submits, redirects to /dashboard"}}
  {{"step_number": 5, "action": "Assert h1 contains text 'Dashboard'", "expected_result": "Dashboard heading visible"}}
  {{"step_number": 6, "action": "Assert URL contains /dashboard", "expected_result": "URL confirms navigation"}}

For "sql" test cases, each step MUST include actual SQL queries:
  {{"step_number": 1, "action": "Execute: SELECT COUNT(*) FROM users WHERE status = 'active'", "expected_result": "Row count >= 1"}}
  {{"step_number": 2, "action": "Execute: SELECT email FROM users WHERE id = 1", "expected_result": "Returns non-null email value"}}

=== GUIDELINES ===
- Infer execution_type from the system description: REST APIs → "api", web UI/forms/pages → "ui", databases/ETL → "sql"
- For web applications, generate a MIX of API tests (backend endpoints) and UI tests (user flows through the browser)
- UI test selectors: prefer #id, then [data-testid=...], then tag.class, then generic CSS
- API tests: include the full endpoint path starting with /
- Make test data realistic but safe (use example.com emails, test passwords, etc.)
- Each test case should be independently executable (no dependencies between test cases)
- If requirements are provided, ensure EACH requirement has at least one corresponding test case
- If BRD/PRD context is provided, derive test scenarios from the business rules and acceptance criteria described

Return ONLY the JSON array, no other text.""")

    user_prompt = "\n".join(user_prompt_parts)

    provider = get_llm_provider()
    messages = [{"role": "user", "content": user_prompt}]

    # Use the smart model for generation (Sonnet instead of Haiku)
    # 16K tokens allows ~10-15 detailed test cases with full steps
    response = provider.complete(
        system=system_prompt,
        messages=messages,
        max_tokens=16384,
        temperature=0.4,
        model=provider.smart_model,
    )

    logger.info(
        "LLM response: model=%s, tokens_in=%d, tokens_out=%d, response_len=%d chars",
        response.model, response.tokens_in, response.tokens_out, len(response.text),
    )

    # Robust JSON parsing — handle markdown fences, truncated responses, trailing commas
    import re as _re

    text = response.text.strip()

    # Step 1: Strip markdown code fences
    md_match = _re.search(r"```(?:json)?\s*\n?(.*?)```", text, _re.DOTALL)
    if md_match:
        text = md_match.group(1).strip()
    elif text.startswith("```"):
        # Truncated response — opening fence but no closing fence
        md_open = _re.match(r"^```(?:json)?\s*\n?", text)
        if md_open:
            text = text[md_open.end():]
            if text.rstrip().endswith("```"):
                text = text.rstrip()[:-3]
            text = text.strip()

    # Step 2: Extract JSON array
    arr_start = text.find("[")
    arr_end = text.rfind("]")
    if arr_start != -1 and arr_end != -1 and arr_end > arr_start:
        text = text[arr_start : arr_end + 1]

    # Step 3: Fix trailing commas (common LLM error)
    text = _re.sub(r",\s*(\]|\})", r"\1", text)

    # Step 4: Try to parse, with truncation repair
    parsed = None
    try:
        parsed = json.loads(text)
    except json.JSONDecodeError as exc:
        logger.warning(
            "JSON parse error at pos %d: %s — attempting truncation repair",
            exc.pos, exc.msg,
        )
        # Try closing open brackets/strings to salvage partial results
        for suffix in ["]", "}", "}]", "\"}]", "\"}]}", "null}]", "\"\"}}]"]:
            try:
                parsed = json.loads(text + suffix)
                logger.info("Fixed truncated JSON with suffix: %s — salvaged %d items",
                            suffix, len(parsed) if isinstance(parsed, list) else 0)
                break
            except json.JSONDecodeError:
                continue

    if not parsed or not isinstance(parsed, list):
        raise ValueError(
            f"LLM did not return a parseable JSON array "
            f"(response length: {len(response.text)} chars, "
            f"tokens: {response.tokens_in}+{response.tokens_out})"
        )

    # Filter out incomplete items (from truncation)
    valid_items = []
    for item in parsed:
        if isinstance(item, dict) and item.get("title") and len(str(item.get("title", ""))) >= 5:
            valid_items.append(item)

    if not valid_items:
        raise ValueError("LLM returned JSON array but no valid test case items")

    provider_name = response.provider or "unknown"
    model_name = response.model or "unknown"
    for item in valid_items:
        item["provider"] = provider_name
        item["model"] = model_name
    track_cost(
        db,
        user_id=user_id,
        project_id=project_id,
        operation_type="llm",
        provider=provider_name,
        model=model_name,
        tokens_in=response.tokens_in,
        tokens_out=response.tokens_out,
    )
    logger.info(
        "LLM generated %d test cases (%s/%s, %d+%d tokens, %d KB entries used)",
        len(valid_items), provider_name, model_name,
        response.tokens_in, response.tokens_out, kb_count,
    )
    # Attach KB metadata so the generate endpoint can track it
    for item in valid_items:
        item["_kb_entries_used"] = kb_count
    return valid_items


def _mock_generate(
    description: str, domain: str, sub_domain: str, count: int
) -> list[dict]:
    """Generate mock test cases as a fallback."""
    results = []
    for i in range(min(count, 20)):
        results.append({
            "title": f"Verify {sub_domain} {domain} functionality - scenario {i + 1}",
            "description": f"Test case generated from: {description[:100]}...",
            "preconditions": f"System is configured for {domain}/{sub_domain}. Test data is loaded.",
            "test_steps": [
                {
                    "step_number": 1,
                    "action": f"Navigate to the {sub_domain} module",
                    "expected_result": f"{sub_domain} module loads successfully",
                },
                {
                    "step_number": 2,
                    "action": "Execute the primary workflow",
                    "expected_result": "Workflow completes without errors",
                },
                {
                    "step_number": 3,
                    "action": "Validate the output data",
                    "expected_result": "Output matches expected results",
                },
            ],
            "expected_result": "All validations pass successfully",
            "priority": ["P1", "P2", "P2", "P3"][i % 4],
            "category": ["functional", "integration", "regression", "smoke", "e2e"][i % 5],
            "domain_tags": [domain, sub_domain],
            "provider": "mock",
            "model": "mock",
            "metadata": {"source": "mock_generator", "version": "1.0"},
        })
    return results


# ---------------------------------------------------------------------------
# GET /{project_id}/test-cases/
# ---------------------------------------------------------------------------
@router.get(
    "/{project_id}/test-cases",
    response_model=list[TestCaseResponse],
    summary="List test cases for a project",
)
def list_test_cases(
    project_id: uuid.UUID,
    test_status: Optional[str] = Query(None, alias="status", description="Filter by status"),
    priority: Optional[str] = Query(None, description="Filter by priority (P1/P2/P3/P4)"),
    category: Optional[str] = Query(None, description="Filter by category"),
    source: Optional[str] = Query(None, description="Filter by source (ai_generated/manual/hybrid)"),
    execution_type: Optional[str] = Query(None, description="Filter by execution type (api/ui/sql/manual)"),
    limit: int = Query(50, ge=1, le=500, description="Page size"),
    offset: int = Query(0, ge=0, description="Offset for pagination"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    List test cases for a project with optional filters and pagination.
    """
    _get_project_or_404(project_id, db)

    query = db.query(TestCase).filter(TestCase.project_id == project_id)

    if test_status:
        query = query.filter(TestCase.status == test_status)
    if priority:
        query = query.filter(TestCase.priority == priority)
    if category:
        query = query.filter(TestCase.category == category)
    if source:
        query = query.filter(TestCase.source == source)
    if execution_type:
        query = query.filter(TestCase.execution_type == execution_type)

    test_cases = (
        query.order_by(TestCase.created_at.desc())
        .offset(offset)
        .limit(limit)
        .all()
    )

    return [TestCaseResponse.model_validate(tc) for tc in test_cases]


# ---------------------------------------------------------------------------
# POST /{project_id}/test-cases/
# ---------------------------------------------------------------------------
@router.post(
    "/{project_id}/test-cases",
    response_model=TestCaseResponse,
    summary="Add a manual test case",
    status_code=status.HTTP_201_CREATED,
)
def create_test_case(
    project_id: uuid.UUID,
    body: TestCaseCreate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Create a test case manually within a project."""
    _get_project_or_404(project_id, db)

    # Check duplicate test_case_id
    existing = db.query(TestCase).filter(
        TestCase.project_id == project_id,
        TestCase.test_case_id == body.test_case_id,
    ).first()
    if existing:
        raise HTTPException(
            status.HTTP_409_CONFLICT,
            detail=f"Test case '{body.test_case_id}' already exists in this project",
        )

    # Validate requirement_id if provided
    if body.requirement_id:
        req = db.query(Requirement).filter(
            Requirement.id == body.requirement_id,
            Requirement.project_id == project_id,
        ).first()
        if req is None:
            raise HTTPException(
                status.HTTP_400_BAD_REQUEST,
                detail="Requirement not found in this project",
            )

    tc = TestCase(
        project_id=project_id,
        requirement_id=body.requirement_id,
        test_case_id=body.test_case_id,
        title=sanitize_string(body.title) or body.title,
        description=sanitize_string(body.description) if body.description else None,
        preconditions=sanitize_string(body.preconditions) if body.preconditions else None,
        test_steps=[s.model_dump() for s in body.test_steps] if body.test_steps else None,
        expected_result=sanitize_string(body.expected_result) if body.expected_result else None,
        test_data=body.test_data,
        priority=body.priority,
        category=body.category,
        domain_tags=body.domain_tags,
        execution_type=body.execution_type,
        source=body.source,
        status="draft",
        created_by=current_user.id,
    )
    db.add(tc)
    db.flush()

    audit_log(
        db,
        user_id=current_user.id,
        action="create_test_case",
        entity_type="test_case",
        entity_id=str(tc.id),
        details={"project_id": str(project_id), "test_case_id": tc.test_case_id},
        ip_address=get_client_ip(request),
    )

    return TestCaseResponse.model_validate(tc)


# ---------------------------------------------------------------------------
# GET /{project_id}/test-cases/upload-template
# ---------------------------------------------------------------------------
@router.get(
    "/{project_id}/test-cases/upload-template",
    summary="Download blank Excel template for bulk upload",
)
def download_upload_template(
    project_id: uuid.UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Returns a pre-formatted Excel template with headers, validation notes,
    and 2 example rows so users know what format to use for bulk upload.
    """
    _get_project_or_404(project_id, db)

    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Cases"

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2E5090", end_color="2E5090", fill_type="solid")
    note_font = Font(italic=True, color="888888", size=9)

    headers = [
        "Test Case ID *",
        "Title *",
        "Description",
        "Preconditions",
        "Priority",
        "Category",
        "Execution Type",
        "Test Steps",
        "Expected Result",
    ]

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Validation notes row
    notes = [
        "Unique ID, e.g. TC-001",
        "Short descriptive title",
        "What does this test validate?",
        "Setup steps needed before test",
        "P1 / P2 / P3 / P4",
        "functional / integration / regression / smoke / e2e",
        "api / ui / sql / manual",
        "Format: 1. Action -> Expected\n2. Action -> Expected",
        "Overall expected outcome",
    ]
    for col_idx, note in enumerate(notes, 1):
        cell = ws.cell(row=2, column=col_idx, value=note)
        cell.font = note_font
        cell.alignment = Alignment(wrap_text=True)

    # Example rows
    examples = [
        [
            "TC-001",
            "Verify user login returns JWT",
            "Test that valid credentials return a JWT access token",
            "Test user exists: test@example.com / pass123",
            "P1",
            "functional",
            "api",
            '1. POST /api/auth/login with {"email":"test@example.com","password":"pass123"} -> Returns 200 with access_token\n2. Decode JWT -> Contains user_id and exp fields',
            "User receives a valid JWT token on successful login",
        ],
        [
            "TC-002",
            "Verify dashboard loads after login",
            "Test that authenticated user can see the dashboard",
            "User is logged in with valid session",
            "P2",
            "ui",
            "ui",
            "1. Navigate to /dashboard -> Page loads with heading 'Dashboard'\n2. Assert sidebar menu has 'Projects' link -> Link is visible and clickable",
            "Dashboard page renders correctly with navigation",
        ],
    ]
    for row_idx, example in enumerate(examples, 3):
        for col_idx, value in enumerate(example, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True)

    # Column widths
    widths = [16, 35, 35, 30, 10, 15, 15, 60, 35]
    for col_idx, width in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    # Freeze top 2 rows
    ws.freeze_panes = "A3"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=test_case_template_{project_id}.xlsx"
        },
    )


# ---------------------------------------------------------------------------
# POST /{project_id}/test-cases/bulk-upload — MUST be before {tc_id} routes
# ---------------------------------------------------------------------------
@router.post(
    "/{project_id}/test-cases/bulk-upload",
    summary="Bulk upload test cases from Excel file",
    status_code=status.HTTP_201_CREATED,
)
async def bulk_upload_test_cases(
    project_id: uuid.UUID,
    request: Request,
    file: UploadFile = File(..., description="Excel (.xlsx) file with test cases"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Parse an uploaded Excel file and create test cases in bulk.

    Expects columns matching the download template:
    Test Case ID, Title, Description, Preconditions, Priority, Category,
    Execution Type, Test Steps, Expected Result.

    Skips the notes row (row 2) and example rows if present.
    Returns a summary of created / skipped / errored rows.
    """
    _get_project_or_404(project_id, db)

    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            detail="Only Excel files (.xlsx) are supported. Download the template first.",
        )

    import openpyxl

    try:
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
    except Exception as exc:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            detail=f"Cannot read Excel file: {exc}",
        )

    # Parse rows (skip header row 1 and notes row 2)
    created = []
    skipped = []
    errors = []

    # Get the next TC number for auto-ID generation
    max_tc = (
        db.query(TestCase.test_case_id)
        .filter(TestCase.project_id == project_id)
        .all()
    )
    existing_ids = {row[0] for row in max_tc}

    VALID_PRIORITIES = {"P1", "P2", "P3", "P4"}
    VALID_CATEGORIES = {"functional", "integration", "regression", "smoke", "e2e"}
    VALID_EXEC_TYPES = {"api", "ui", "sql", "manual"}

    for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        # Skip completely empty rows
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        # Pad row to 9 columns
        cells = list(row) + [None] * max(0, 9 - len(row))
        tc_id = str(cells[0] or "").strip()
        title = str(cells[1] or "").strip()

        # Skip notes/example hint rows
        if tc_id.lower().startswith("unique id") or title.lower().startswith("short descriptive"):
            continue

        # Validate required fields
        if not tc_id or not title:
            errors.append({"row": row_idx, "error": "Missing Test Case ID or Title"})
            continue

        if tc_id in existing_ids:
            skipped.append({"row": row_idx, "tc_id": tc_id, "reason": "Duplicate ID"})
            continue

        description = str(cells[2] or "").strip() or None
        preconditions = str(cells[3] or "").strip() or None
        priority = str(cells[4] or "P2").strip().upper()
        category = str(cells[5] or "functional").strip().lower()
        exec_type = str(cells[6] or "manual").strip().lower()
        steps_text = str(cells[7] or "").strip()
        expected_result = str(cells[8] or "").strip() or None

        # Validate enum values
        if priority not in VALID_PRIORITIES:
            priority = "P2"
        if category not in VALID_CATEGORIES:
            category = "functional"
        if exec_type not in VALID_EXEC_TYPES:
            exec_type = "manual"

        # Parse test steps from text format: "1. Action -> Expected\n2. ..."
        test_steps = None
        if steps_text:
            test_steps = []
            import re
            step_lines = re.split(r"\n+", steps_text)
            for snum, line in enumerate(step_lines, 1):
                line = line.strip()
                if not line:
                    continue
                # Remove leading step number like "1." or "1)"
                line = re.sub(r"^\d+[\.\)]\s*", "", line)
                if " -> " in line:
                    action, exp = line.split(" -> ", 1)
                else:
                    action, exp = line, ""
                test_steps.append({
                    "step_number": snum,
                    "action": action.strip(),
                    "expected_result": exp.strip(),
                })

        try:
            tc = TestCase(
                project_id=project_id,
                test_case_id=tc_id,
                title=sanitize_string(title) or title,
                description=sanitize_string(description) if description else None,
                preconditions=sanitize_string(preconditions) if preconditions else None,
                test_steps=test_steps,
                expected_result=sanitize_string(expected_result) if expected_result else None,
                priority=priority,
                category=category,
                execution_type=exec_type,
                source="manual",
                status="draft",
                created_by=current_user.id,
            )
            db.add(tc)
            db.flush()
            existing_ids.add(tc_id)
            created.append({"row": row_idx, "tc_id": tc_id, "title": title})
        except Exception as exc:
            db.rollback()
            errors.append({"row": row_idx, "tc_id": tc_id, "error": str(exc)})

    # Commit all successful creates
    if created:
        db.commit()

    audit_log(
        db,
        user_id=current_user.id,
        action="bulk_upload_test_cases",
        entity_type="test_case",
        entity_id=str(project_id),
        details={
            "created": len(created),
            "skipped": len(skipped),
            "errors": len(errors),
            "filename": file.filename,
        },
        ip_address=get_client_ip(request),
    )

    return {
        "created": len(created),
        "skipped": len(skipped),
        "errors": len(errors),
        "details": {
            "created": created,
            "skipped": skipped,
            "errors": errors,
        },
    }


# ---------------------------------------------------------------------------
# GET /{project_id}/test-cases/{tc_id}
# ---------------------------------------------------------------------------
@router.get(
    "/{project_id}/test-cases/{tc_id}",
    response_model=TestCaseResponse,
    summary="Get a single test case",
)
def get_test_case(
    project_id: uuid.UUID,
    tc_id: uuid.UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Return a single test case by ID."""
    _get_project_or_404(project_id, db)
    tc = _get_test_case_or_404(tc_id, project_id, db)
    return TestCaseResponse.model_validate(tc)


# ---------------------------------------------------------------------------
# PUT /{project_id}/test-cases/{tc_id}
# ---------------------------------------------------------------------------
@router.put(
    "/{project_id}/test-cases/{tc_id}",
    response_model=TestCaseResponse,
    summary="Update a test case",
)
def update_test_case(
    project_id: uuid.UUID,
    tc_id: uuid.UUID,
    body: TestCaseUpdate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Update fields of an existing test case."""
    _get_project_or_404(project_id, db)
    tc = _get_test_case_or_404(tc_id, project_id, db)

    if body.title is not None:
        tc.title = sanitize_string(body.title) or body.title
    if body.description is not None:
        tc.description = sanitize_string(body.description)
    if body.preconditions is not None:
        tc.preconditions = sanitize_string(body.preconditions)
    if body.test_steps is not None:
        tc.test_steps = [s.model_dump() for s in body.test_steps]
    if body.expected_result is not None:
        tc.expected_result = sanitize_string(body.expected_result)
    if body.test_data is not None:
        tc.test_data = body.test_data
    if body.priority is not None:
        tc.priority = body.priority
    if body.category is not None:
        tc.category = body.category
    if body.domain_tags is not None:
        tc.domain_tags = body.domain_tags
    if body.execution_type is not None:
        tc.execution_type = body.execution_type
    if body.status is not None:
        tc.status = body.status

    db.flush()

    audit_log(
        db,
        user_id=current_user.id,
        action="update_test_case",
        entity_type="test_case",
        entity_id=str(tc.id),
        details=body.model_dump(exclude_none=True),
        ip_address=get_client_ip(request),
    )

    return TestCaseResponse.model_validate(tc)


# ---------------------------------------------------------------------------
# DELETE /{project_id}/test-cases/{tc_id}
# ---------------------------------------------------------------------------
@router.delete(
    "/{project_id}/test-cases/{tc_id}",
    response_model=MessageResponse,
    summary="Delete a test case",
)
def delete_test_case(
    project_id: uuid.UUID,
    tc_id: uuid.UUID,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Delete a test case from a project."""
    _get_project_or_404(project_id, db)
    tc = _get_test_case_or_404(tc_id, project_id, db)

    tc_display = tc.test_case_id
    db.delete(tc)
    db.flush()

    audit_log(
        db,
        user_id=current_user.id,
        action="delete_test_case",
        entity_type="test_case",
        entity_id=str(tc_id),
        details={"project_id": str(project_id), "test_case_id": tc_display},
        ip_address=get_client_ip(request),
    )

    return MessageResponse(message=f"Test case '{tc_display}' deleted")


# ---------------------------------------------------------------------------
# POST /{project_id}/test-cases/bulk-delete
# ---------------------------------------------------------------------------
@router.post(
    "/{project_id}/test-cases/bulk-delete",
    response_model=MessageResponse,
    summary="Bulk delete test cases by IDs",
)
def bulk_delete_test_cases(
    project_id: uuid.UUID,
    body: dict,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Delete multiple test cases at once (used for rejecting generated TCs)."""
    _get_project_or_404(project_id, db)

    tc_ids = body.get("test_case_ids", [])
    if not tc_ids:
        raise HTTPException(status_code=400, detail="No test_case_ids provided")
    if len(tc_ids) > 200:
        raise HTTPException(status_code=400, detail="Maximum 200 test cases per bulk delete")

    # Convert to UUIDs
    try:
        tc_uuids = [uuid.UUID(str(tid)) for tid in tc_ids]
    except (ValueError, AttributeError):
        raise HTTPException(status_code=400, detail="Invalid test case ID format")

    deleted_count = (
        db.query(TestCase)
        .filter(
            TestCase.id.in_(tc_uuids),
            TestCase.project_id == project_id,
        )
        .delete(synchronize_session="fetch")
    )
    db.flush()

    audit_log(
        db,
        user_id=current_user.id,
        action="bulk_delete_test_cases",
        entity_type="test_case",
        entity_id=str(project_id),
        details={"requested": len(tc_ids), "deleted": deleted_count},
        ip_address=get_client_ip(request),
    )

    logger.info("Bulk deleted %d test cases for project %s", deleted_count, project_id)

    return MessageResponse(
        message=f"Deleted {deleted_count} test cases",
        detail=f"Requested: {len(tc_ids)}, deleted: {deleted_count}",
    )


# ---------------------------------------------------------------------------
# POST /{project_id}/test-cases/{tc_id}/rate
# ---------------------------------------------------------------------------
@router.post(
    "/{project_id}/test-cases/{tc_id}/rate",
    response_model=TestCaseResponse,
    summary="Rate a test case",
)
def rate_test_case(
    project_id: uuid.UUID,
    tc_id: uuid.UUID,
    body: TestCaseRateRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Rate a test case (1-5 stars) with optional feedback text.

    Creates a FeedbackEntry and updates the test case's rating.
    """
    _get_project_or_404(project_id, db)
    tc = _get_test_case_or_404(tc_id, project_id, db)

    # Create feedback entry
    feedback = FeedbackEntry(
        test_case_id=tc.id,
        rating=body.rating,
        original_content={
            "title": tc.title,
            "description": tc.description,
            "test_steps": tc.test_steps,
        },
        feedback_text=sanitize_string(body.feedback_text) if body.feedback_text else None,
        feedback_type="quality",
        created_by=current_user.id,
    )
    db.add(feedback)

    # Update test case rating (keep the latest rating)
    tc.rating = body.rating
    tc.rating_feedback = sanitize_string(body.feedback_text) if body.feedback_text else None

    db.flush()

    audit_log(
        db,
        user_id=current_user.id,
        action="rate_test_case",
        entity_type="test_case",
        entity_id=str(tc.id),
        details={"rating": body.rating, "test_case_id": tc.test_case_id},
        ip_address=get_client_ip(request),
    )

    return TestCaseResponse.model_validate(tc)


# ---------------------------------------------------------------------------
# POST /{project_id}/test-cases/export
# ---------------------------------------------------------------------------
@router.post(
    "/{project_id}/test-cases/export",
    summary="Export test cases to Excel",
)
def export_test_cases(
    project_id: uuid.UUID,
    body: TestCaseExportRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Export selected test cases to an Excel file.

    Returns the file as a downloadable attachment.
    """
    _get_project_or_404(project_id, db)

    # Fetch requested test cases
    test_cases = (
        db.query(TestCase)
        .filter(
            TestCase.id.in_(body.test_case_ids),
            TestCase.project_id == project_id,
        )
        .order_by(TestCase.test_case_id.asc())
        .all()
    )

    if not test_cases:
        raise HTTPException(
            status.HTTP_404_NOT_FOUND,
            detail="No test cases found for the given IDs",
        )

    # Try openpyxl for Excel generation
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test Cases"

        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2E5090", end_color="2E5090", fill_type="solid")

        # Define columns
        headers = [
            "Test Case ID",
            "Title",
            "Description",
            "Preconditions",
            "Priority",
            "Category",
            "Status",
            "Source",
        ]
        if body.include_steps:
            headers.append("Test Steps")
        headers.append("Expected Result")
        if body.include_test_data:
            headers.append("Test Data")

        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Write data rows
        for row_idx, tc in enumerate(test_cases, 2):
            col = 1
            ws.cell(row=row_idx, column=col, value=tc.test_case_id); col += 1
            ws.cell(row=row_idx, column=col, value=tc.title); col += 1
            ws.cell(row=row_idx, column=col, value=tc.description or ""); col += 1
            ws.cell(row=row_idx, column=col, value=tc.preconditions or ""); col += 1
            ws.cell(row=row_idx, column=col, value=tc.priority); col += 1
            ws.cell(row=row_idx, column=col, value=tc.category); col += 1
            ws.cell(row=row_idx, column=col, value=tc.status); col += 1
            ws.cell(row=row_idx, column=col, value=tc.source); col += 1

            if body.include_steps and tc.test_steps:
                steps_text = "\n".join(
                    f"{s.get('step_number', i+1)}. {s.get('action', '')} -> {s.get('expected_result', '')}"
                    for i, s in enumerate(tc.test_steps)
                )
                cell = ws.cell(row=row_idx, column=col, value=steps_text)
                cell.alignment = Alignment(wrap_text=True)
                col += 1
            elif body.include_steps:
                ws.cell(row=row_idx, column=col, value=""); col += 1

            ws.cell(row=row_idx, column=col, value=tc.expected_result or ""); col += 1

            if body.include_test_data:
                ws.cell(
                    row=row_idx,
                    column=col,
                    value=json.dumps(tc.test_data, indent=2) if tc.test_data else "",
                )
                col += 1

        # Auto-width columns
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        audit_log(
            db,
            user_id=current_user.id,
            action="export_test_cases",
            entity_type="test_case",
            entity_id=str(project_id),
            details={"count": len(test_cases), "format": body.format},
            ip_address=get_client_ip(request),
        )

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=test_cases_{project_id}.xlsx"
            },
        )

    except ImportError:
        # Fallback to CSV if openpyxl not installed
        logger.warning("openpyxl not installed; falling back to CSV export")
        return _export_csv(test_cases, project_id, body, current_user, request, db)


def _export_csv(
    test_cases: list[TestCase],
    project_id: uuid.UUID,
    body: TestCaseExportRequest,
    current_user: User,
    request: Request,
    db: Session,
) -> StreamingResponse:
    """Fallback CSV export when openpyxl is not available."""
    import csv

    buffer = io.StringIO()
    writer = csv.writer(buffer)

    headers = [
        "Test Case ID", "Title", "Description", "Preconditions",
        "Priority", "Category", "Status", "Source", "Expected Result",
    ]
    writer.writerow(headers)

    for tc in test_cases:
        writer.writerow([
            tc.test_case_id,
            tc.title,
            tc.description or "",
            tc.preconditions or "",
            tc.priority,
            tc.category,
            tc.status,
            tc.source,
            tc.expected_result or "",
        ])

    audit_log(
        db,
        user_id=current_user.id,
        action="export_test_cases",
        entity_type="test_case",
        entity_id=str(project_id),
        details={"count": len(test_cases), "format": "csv"},
        ip_address=get_client_ip(request),
    )

    content = buffer.getvalue().encode("utf-8")
    return StreamingResponse(
        io.BytesIO(content),
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename=test_cases_{project_id}.csv"
        },
    )


# ---------------------------------------------------------------------------
# POST /chat  (Feature 6: Chat-Based Test Generation Agent)
# ---------------------------------------------------------------------------
@router.post(
    "/{project_id}/test-cases/chat",
    summary="Chat-based test case generation",
)
def chat_generate(
    project_id: uuid.UUID,
    body: ChatGenerateRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Multi-turn conversational test generation.

    The agent reviews project context (app_profile, requirements, existing TCs)
    and decides whether to ask clarifying questions or generate test cases.

    Returns:
    - action="question": agent needs more info (message contains the question)
    - action="confirm": agent proposes a config before generating
    - action="generate": agent generated test cases (test_cases field populated)
    """
    from core.llm_provider import get_llm_provider

    project = _get_project_or_404(project_id, db)

    # Build project context summary
    app_profile = project.app_profile or {}
    requirements = (
        db.query(Requirement)
        .filter(Requirement.project_id == project_id)
        .all()
    )

    tc_count = (
        db.query(func.count(TestCase.id))
        .filter(TestCase.project_id == project_id)
        .scalar()
    ) or 0

    # Build context for the agent
    context_parts = []

    context_parts.append(f"Project: {project.name} (domain: {project.domain}, sub-domain: {project.sub_domain})")
    if project.description:
        context_parts.append(f"Description: {project.description[:500]}")
    context_parts.append(f"Existing test cases: {tc_count}")
    context_parts.append(f"Requirements: {len(requirements)}")

    if requirements:
        req_summary = []
        for r in requirements[:15]:
            req_summary.append(f"  [{r.priority}] {r.req_id}: {r.title}")
        context_parts.append("Requirements list:\n" + "\n".join(req_summary))

    if app_profile:
        ap_summary = []
        if app_profile.get("app_url"):
            ap_summary.append(f"App URL: {app_profile['app_url']}")
        if app_profile.get("api_base_url"):
            ap_summary.append(f"API Base: {app_profile['api_base_url']}")
        endpoints = app_profile.get("api_endpoints", [])
        if endpoints:
            ap_summary.append(f"API Endpoints: {len(endpoints)} configured")
            for ep in endpoints[:10]:
                if isinstance(ep, dict):
                    ap_summary.append(f"  {ep.get('method', 'GET')} {ep.get('path', '')}")
        ui_pages = app_profile.get("ui_pages", [])
        if ui_pages:
            ap_summary.append(f"UI Pages: {len(ui_pages)} configured")
        if app_profile.get("auth"):
            ap_summary.append("Auth: configured")
        context_parts.append("App Profile:\n" + "\n".join(ap_summary))

    project_context = "\n".join(context_parts)

    # System prompt for the chat agent
    system_prompt = f"""You are an AI QA test generation assistant for the QAForge platform.
Your job is to help users create high-quality, execution-ready test cases.

You have access to this project's context:

{project_context}

== YOUR WORKFLOW ==

1. EVALUATE if you have enough context to generate good test cases. Consider:
   - Do you know WHAT features/endpoints/pages to test?
   - Do you know the execution type (api, ui, sql)?
   - Are the user's requirements clear and specific?
   - Is there enough detail to write concrete, executable test steps?

2. If context is INCOMPLETE (first 1-2 turns), ask 1-2 focused clarifying questions.
   Good questions:
   - "Which specific features would you like to test? For example: authentication, CRUD operations, dashboard, etc."
   - "Should I generate API tests, UI tests, or a mix of both?"
   - "How many test cases would you like? And what priority level?"
   - "I notice your app profile has 29 API endpoints. Should I focus on specific ones?"

3. If you have ENOUGH context, respond with a JSON block to trigger generation:
   ```json
   {{"action": "generate", "config": {{"description": "...", "count": 10, "execution_type": "api", "priority": "P1"}}}}
   ```

RULES:
- Ask a MAXIMUM of 2 rounds of questions, then generate with best available context
- Be conversational and helpful, not formal
- Reference the project's actual endpoints, pages, and requirements in your questions
- When ready to generate, ALWAYS include the JSON config block
- Never generate test cases directly in chat — use the JSON config block to trigger the generation pipeline
"""

    # Build messages for LLM
    llm_messages = []
    for msg in body.messages:
        llm_messages.append({"role": msg.role, "content": msg.content})

    # Call LLM (use fast model for chat, it's cheaper)
    provider = get_llm_provider()
    try:
        result = provider.complete(
            system=system_prompt,
            messages=llm_messages,
            model=getattr(provider, "fast_model", None) or getattr(provider, "model", None),
            max_tokens=2048,
            temperature=0.5,
        )
    except Exception as exc:
        logger.error("Chat generation LLM call failed: %s", exc, exc_info=True)
        return ChatGenerateResponse(
            message=ChatMessage(
                role="assistant",
                content="I'm having trouble connecting to the AI model. Please try again in a moment.",
            ),
            action="question",
        )

    response_text = result.text.strip()

    # Check if the response contains a generation config JSON
    import re as _re
    json_match = _re.search(r'\{["\s]*"action"["\s]*:\s*"generate".*?\}', response_text, _re.DOTALL)
    if json_match:
        try:
            config_text = json_match.group(0)
            # Find the complete JSON block (handle nested braces)
            start = response_text.find(config_text)
            depth = 0
            end = start
            for i, ch in enumerate(response_text[start:], start):
                if ch == "{":
                    depth += 1
                elif ch == "}":
                    depth -= 1
                    if depth == 0:
                        end = i
                        break
            config_block = response_text[start:end + 1]
            config = json.loads(config_block)

            if config.get("action") == "generate":
                gen_config = config.get("config", {})
                description = gen_config.get("description", body.messages[0].content if body.messages else "")
                count = gen_config.get("count", 10)
                execution_type = gen_config.get("execution_type", body.execution_type)
                priority = gen_config.get("priority")

                # Build requirement context
                requirement_context = ""
                req_ids = body.requirement_ids
                if req_ids:
                    sel_reqs = [r for r in requirements if r.id in req_ids]
                else:
                    sel_reqs = requirements

                if sel_reqs:
                    req_lines = []
                    for r in sel_reqs[:20]:
                        req_lines.append(f"[{r.priority}] {r.req_id}: {r.title}\n  {r.description or 'N/A'}")
                    requirement_context = "\n".join(req_lines)

                # Generate test cases using the existing pipeline
                try:
                    generated = _generate_via_llm(
                        description=description[:4000],
                        domain=project.domain,
                        sub_domain=project.sub_domain,
                        requirement_context=requirement_context,
                        count=min(count, 50),
                        additional_context=None,
                        project_id=project_id,
                        user_id=current_user.id,
                        db=db,
                        app_profile=app_profile,
                    )

                    # Save generated test cases to DB
                    saved_tcs = []
                    tc_seq = (
                        db.query(func.count(TestCase.id))
                        .filter(TestCase.project_id == project_id)
                        .scalar()
                    ) or 0

                    for i, tc_data in enumerate(generated):
                        tc_seq += 1
                        tc = TestCase(
                            project_id=project_id,
                            test_case_id=tc_data.get("test_case_id", f"TC-{tc_seq:04d}"),
                            title=tc_data.get("title", f"Test Case {tc_seq}"),
                            description=tc_data.get("description"),
                            preconditions=tc_data.get("preconditions"),
                            test_steps=tc_data.get("test_steps", []),
                            expected_result=tc_data.get("expected_result"),
                            test_data=tc_data.get("test_data"),
                            priority=tc_data.get("priority", priority or "P2"),
                            category=tc_data.get("category", "functional"),
                            domain_tags=tc_data.get("domain_tags", []),
                            execution_type=tc_data.get("execution_type", execution_type or "api"),
                            source="ai_generated",
                            created_by=current_user.id,
                        )

                        # Link to requirement if possible
                        if req_ids and len(req_ids) == 1:
                            tc.requirement_id = req_ids[0]
                        elif tc_data.get("requirement_id"):
                            tc.requirement_id = tc_data["requirement_id"]

                        db.add(tc)
                        db.flush()
                        saved_tcs.append(TestCaseResponse.model_validate(tc))

                    audit_log(
                        db,
                        user_id=current_user.id,
                        action="chat_generate_test_cases",
                        entity_type="test_case",
                        entity_id=str(project_id),
                        details={"count": len(saved_tcs), "source": "chat_agent"},
                        ip_address=get_client_ip(request),
                    )

                    # Remove the JSON config from the response text for cleaner message
                    clean_text = response_text[:start].strip()
                    if not clean_text:
                        clean_text = f"Generated {len(saved_tcs)} test cases based on our conversation."

                    return ChatGenerateResponse(
                        message=ChatMessage(role="assistant", content=clean_text),
                        action="generate",
                        test_cases=saved_tcs,
                        suggested_config=gen_config,
                    )

                except Exception as exc:
                    logger.error("Chat generation pipeline failed: %s", exc, exc_info=True)
                    return ChatGenerateResponse(
                        message=ChatMessage(
                            role="assistant",
                            content=f"I tried to generate test cases but encountered an error: {str(exc)[:200]}. Could you try again?",
                        ),
                        action="question",
                    )

        except (json.JSONDecodeError, ValueError):
            pass  # Not valid JSON — treat as a regular question response

    # No generation config found — this is a question/clarification
    return ChatGenerateResponse(
        message=ChatMessage(role="assistant", content=response_text),
        action="question",
    )
