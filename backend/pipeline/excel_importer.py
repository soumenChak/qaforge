"""
QAForge -- Excel Reference Test Case Importer
===============================================
Parses Excel test case files (MDM and DE formats) into QAForge test case
dicts that can be:
  1. Imported directly as reference test cases via the Agent API
  2. Serialised as text context for LLM prompt injection (reference_tc_context)

Supports two auto-detected formats:
  - MDM format:  Test case ID | Test Category | User Story | Description | Pre-Requisites | Step | Test Steps | Expected Result
  - DE format:   TC# | Short Name | Type | Description | Expected Result | Prerequisite | Step# | Step Description | Expected Results | SQL Scripts

Both formats group multi-row test cases by the first non-None ID column.
"""

from __future__ import annotations

import json
import logging
import re
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

logger = logging.getLogger(__name__)

try:
    import openpyxl
except ImportError:
    openpyxl = None  # type: ignore
    logger.warning("openpyxl not installed — Excel import disabled")


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def parse_excel(
    file_bytes: bytes,
    filename: str = "unknown.xlsx",
) -> Tuple[List[Dict[str, Any]], str]:
    """
    Parse an Excel file into QAForge test case dicts + LLM context text.

    Args:
        file_bytes: Raw bytes of the Excel file.
        filename: Original filename (used for logging/tagging).

    Returns:
        Tuple of (test_cases, reference_text) where:
          - test_cases is a list of QAForge-format test case dicts
          - reference_text is a text summary for LLM prompt injection
    """
    if openpyxl is None:
        raise ImportError("openpyxl is required for Excel parsing. pip install openpyxl")

    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True)
    test_cases: List[Dict[str, Any]] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        # Auto-detect format from header row
        header = [str(c).strip().lower() if c else "" for c in rows[0]]

        fmt = _detect_format(header, sheet_name)
        if fmt == "mdm":
            parsed = _parse_mdm_format(rows, header)
        elif fmt == "de":
            parsed = _parse_de_format(rows, header)
        elif fmt == "requirement":
            # Requirement sheets contain BRD text, not test cases
            continue
        else:
            # Try generic multi-step format
            parsed = _parse_generic_format(rows, header)

        if parsed:
            logger.info("Parsed %d test cases from sheet '%s' (format=%s)", len(parsed), sheet_name, fmt)
            test_cases.extend(parsed)

    wb.close()

    # Generate reference text for LLM context
    reference_text = _build_reference_text(test_cases, filename)

    return test_cases, reference_text


def extract_brd_text(file_bytes: bytes) -> str:
    """
    Extract BRD/requirement text from an Excel file.

    Looks for sheets named 'Req', 'Requirement', 'Requirements', 'BRD'
    and extracts text content.
    """
    if openpyxl is None:
        raise ImportError("openpyxl required")

    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True)
    brd_parts: List[str] = []

    req_sheets = ["req", "requirement", "requirements", "brd", "prd", "spec"]
    for sheet_name in wb.sheetnames:
        if sheet_name.lower().strip() in req_sheets:
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell and str(cell).strip():
                        brd_parts.append(str(cell).strip())

    wb.close()
    return "\n\n".join(brd_parts)


# ---------------------------------------------------------------------------
# Format detection
# ---------------------------------------------------------------------------

def _detect_format(header: List[str], sheet_name: str) -> str:
    """Detect the Excel format from header columns and sheet name."""
    header_text = " ".join(header)

    # Requirement-only sheets
    if sheet_name.lower().strip() in ("req", "requirement", "requirements", "brd", "prd"):
        return "requirement"

    # MDM format: has "test category" and "user story" columns
    if "test category" in header_text and "user story" in header_text:
        return "mdm"
    if "test case id" in header_text and "pre-requisites" in header_text:
        return "mdm"

    # DE format: has "sql scripts" or "test case short name" columns
    if "sql scripts" in header_text:
        return "de"
    if "test case short name" in header_text:
        return "de"
    if "test case #" in header_text or "test case #" in header_text:
        return "de"

    # Check for numbered TC columns
    if header and header[0] and re.match(r"(test case|tc)\s*#?$", header[0]):
        return "de"

    return "generic"


# ---------------------------------------------------------------------------
# MDM format parser
# ---------------------------------------------------------------------------

def _parse_mdm_format(rows: List[tuple], header: List[str]) -> List[Dict[str, Any]]:
    """
    Parse MDM format:
    Test case ID | Test Category | User Story | Description | Pre-Requisites | Step | Test Steps | Expected Result
    """
    # Map column positions
    col_map = _build_col_map(header, {
        "id": ["test case id", "test_case_id", "tc id", "tc_id"],
        "category": ["test category", "category"],
        "user_story": ["user story", "user_story", "requirement"],
        "description": ["description", "test description"],
        "prerequisites": ["pre-requisites", "prerequisites", "preconditions"],
        "step_num": ["step", "step#", "step #", "step_number"],
        "step_action": ["test steps", "step description", "test step", "steps"],
        "step_expected": ["expected result", "expected results", "expected_result"],
    })

    test_cases: List[Dict[str, Any]] = []
    current_tc: Optional[Dict[str, Any]] = None
    step_num = 0

    for row in rows[1:]:  # skip header
        row_id = _cell(row, col_map.get("id"))

        if row_id:
            # New test case starts
            if current_tc:
                test_cases.append(current_tc)

            step_num = 0
            current_tc = {
                "test_case_id": str(row_id).strip(),
                "title": str(_cell(row, col_map.get("description")) or row_id).strip()[:200],
                "description": str(_cell(row, col_map.get("description")) or "").strip(),
                "category": str(_cell(row, col_map.get("category")) or "Functional").strip(),
                "preconditions": str(_cell(row, col_map.get("prerequisites")) or "None").strip(),
                "test_steps": [],
                "expected_result": "",
                "priority": "High",
                "domain_tags": ["MDM"],
                "execution_type": "mdm",
                "test_data": {},
            }

            user_story = _cell(row, col_map.get("user_story"))
            if user_story:
                current_tc["test_data"]["user_story"] = str(user_story).strip()

        if current_tc is None:
            continue

        # Add step from this row
        action = _cell(row, col_map.get("step_action"))
        expected = _cell(row, col_map.get("step_expected"))
        step_number = _cell(row, col_map.get("step_num"))

        if action:
            step_num += 1
            step = {
                "step_number": int(step_number) if step_number else step_num,
                "action": str(action).strip(),
                "expected_result": str(expected).strip() if expected else "",
            }

            # Infer step_type from action text
            action_lower = str(action).lower()
            if "snowflake" in action_lower or "query" in action_lower and "table" in action_lower:
                step["step_type"] = "snowflake"
                step["system"] = "Snowflake"
            elif "reltio" in action_lower or "advanced search" in action_lower:
                step["step_type"] = "reltio_ui"
                step["system"] = "Reltio"
            elif "spreadsheet" in action_lower or "refer" in action_lower:
                step["step_type"] = "manual"
                step["system"] = "Spreadsheet"
            elif "compare" in action_lower or "verify" in action_lower or "ensure" in action_lower:
                step["step_type"] = "data_validation"
                step["system"] = "Cross-system"
            elif "login" in action_lower:
                if "reltio" in action_lower:
                    step["step_type"] = "reltio_ui"
                    step["system"] = "Reltio"
                elif "snowflake" in action_lower:
                    step["step_type"] = "snowflake"
                    step["system"] = "Snowflake"
                else:
                    step["step_type"] = "manual"
            else:
                step["step_type"] = "manual"

            current_tc["test_steps"].append(step)

            # Use last step's expected result as overall expected result
            if expected:
                current_tc["expected_result"] = str(expected).strip()

    # Don't forget the last TC
    if current_tc:
        test_cases.append(current_tc)

    return test_cases


# ---------------------------------------------------------------------------
# DE format parser
# ---------------------------------------------------------------------------

def _parse_de_format(rows: List[tuple], header: List[str]) -> List[Dict[str, Any]]:
    """
    Parse DE format:
    TC# | Short Name | Type | Description | Expected Result | Prerequisite | Step# | Step Description | Expected Results | SQL Scripts
    """
    col_map = _build_col_map(header, {
        "id": ["test case #", "tc#", "tc-#", " test case #", "test case id"],
        "short_name": ["test case short name", "short name", "name"],
        "type": ["type of\ntest case", "type", "test type"],
        "description": ["description"],
        "overall_expected": ["expected test case result", "expected result"],
        "prerequisites": ["prerequisite", "prerequisites", "preconditions", "pre-requisites"],
        "step_num": ["step #", "step#", "step"],
        "step_action": ["step description", "test steps"],
        "step_expected": ["expected results (for step)", "expected results", "expected result (for step)"],
        "sql_scripts": ["sql scripts", "sql", "query"],
    })

    test_cases: List[Dict[str, Any]] = []
    current_tc: Optional[Dict[str, Any]] = None
    step_num = 0

    for row in rows[1:]:  # skip header
        row_id = _cell(row, col_map.get("id"))

        if row_id:
            # New test case starts
            if current_tc:
                test_cases.append(current_tc)

            step_num = 0
            short_name = _cell(row, col_map.get("short_name"))
            description = _cell(row, col_map.get("description"))
            overall_expected = _cell(row, col_map.get("overall_expected"))

            current_tc = {
                "test_case_id": str(row_id).strip(),
                "title": str(short_name or description or row_id).strip()[:200],
                "description": str(description or "").strip(),
                "category": "Migration",
                "preconditions": str(_cell(row, col_map.get("prerequisites")) or "None").strip(),
                "test_steps": [],
                "expected_result": str(overall_expected or "").strip(),
                "priority": "High",
                "domain_tags": ["Data Engineering", "Migration"],
                "execution_type": "sql",
                "test_data": {"sql_scripts": []},
            }

            tc_type = _cell(row, col_map.get("type"))
            if tc_type:
                current_tc["test_data"]["test_type"] = str(tc_type).strip()

        if current_tc is None:
            continue

        # Add step from this row
        action = _cell(row, col_map.get("step_action"))
        expected = _cell(row, col_map.get("step_expected"))
        step_number = _cell(row, col_map.get("step_num"))
        sql_script = _cell(row, col_map.get("sql_scripts"))

        if action:
            step_num += 1
            step: Dict[str, Any] = {
                "step_number": int(step_number) if step_number else step_num,
                "action": str(action).strip(),
                "expected_result": str(expected).strip() if expected else "",
            }

            # Infer step_type
            action_lower = str(action).lower()
            if sql_script:
                sql_text = str(sql_script).strip()
                step["sql_script"] = sql_text
                step["step_type"] = "sql"
                step["system"] = "Database"
                # Track SQL scripts in test_data too
                current_tc["test_data"]["sql_scripts"].append(sql_text)
            elif "source" in action_lower or "oracle" in action_lower:
                step["step_type"] = "oracle"
                step["system"] = "Oracle"
            elif "target" in action_lower or "databricks" in action_lower:
                step["step_type"] = "databricks"
                step["system"] = "Databricks"
            elif "snowflake" in action_lower:
                step["step_type"] = "snowflake"
                step["system"] = "Snowflake"
            elif "compare" in action_lower or "verify" in action_lower:
                step["step_type"] = "data_validation"
                step["system"] = "Cross-system"
            else:
                step["step_type"] = "manual"

            current_tc["test_steps"].append(step)

    if current_tc:
        test_cases.append(current_tc)

    return test_cases


# ---------------------------------------------------------------------------
# Generic format parser (fallback)
# ---------------------------------------------------------------------------

def _parse_generic_format(rows: List[tuple], header: List[str]) -> List[Dict[str, Any]]:
    """Fallback parser for unrecognised but structured test case sheets."""
    # Try to find columns that look like ID, description, steps
    col_map = _build_col_map(header, {
        "id": ["id", "test id", "tc", "test case", "#"],
        "title": ["title", "name", "test name", "short name"],
        "description": ["description", "desc"],
        "step_action": ["steps", "test steps", "step description", "action"],
        "step_expected": ["expected", "expected result", "expected results"],
    })

    if not col_map.get("id"):
        return []

    test_cases: List[Dict[str, Any]] = []
    current_tc: Optional[Dict[str, Any]] = None
    step_num = 0

    for row in rows[1:]:
        row_id = _cell(row, col_map.get("id"))

        if row_id:
            if current_tc:
                test_cases.append(current_tc)
            step_num = 0
            current_tc = {
                "test_case_id": str(row_id).strip(),
                "title": str(_cell(row, col_map.get("title")) or row_id).strip()[:200],
                "description": str(_cell(row, col_map.get("description")) or "").strip(),
                "category": "Functional",
                "preconditions": "None",
                "test_steps": [],
                "expected_result": "",
                "priority": "Medium",
                "domain_tags": [],
                "execution_type": "manual",
                "test_data": {},
            }

        if current_tc is None:
            continue

        action = _cell(row, col_map.get("step_action"))
        expected = _cell(row, col_map.get("step_expected"))

        if action:
            step_num += 1
            current_tc["test_steps"].append({
                "step_number": step_num,
                "action": str(action).strip(),
                "expected_result": str(expected).strip() if expected else "",
                "step_type": "manual",
            })
            if expected:
                current_tc["expected_result"] = str(expected).strip()

    if current_tc:
        test_cases.append(current_tc)

    return test_cases


# ---------------------------------------------------------------------------
# LLM context text builder
# ---------------------------------------------------------------------------

def _build_reference_text(
    test_cases: List[Dict[str, Any]], filename: str
) -> str:
    """
    Build a text summary of parsed test cases for LLM prompt injection.

    This text is injected as reference_tc_context so the LLM generates
    test cases matching this style and detail level.
    """
    parts = [f"Reference test cases from: {filename}\n"]

    for tc in test_cases[:5]:  # Limit to 5 examples to stay within token budget
        parts.append(f"\n--- {tc['test_case_id']}: {tc['title']} ---")
        parts.append(f"Category: {tc.get('category', 'N/A')}")
        parts.append(f"Preconditions: {tc.get('preconditions', 'None')}")
        parts.append(f"Execution type: {tc.get('execution_type', 'manual')}")

        for step in tc.get("test_steps", [])[:6]:  # Limit steps shown
            step_info = f"  Step {step['step_number']}: {step['action']}"
            if step.get("step_type"):
                step_info += f" [{step['step_type']}]"
            if step.get("sql_script"):
                step_info += f"\n    SQL: {step['sql_script'][:150]}"
            step_info += f"\n    Expected: {step.get('expected_result', '')}"
            parts.append(step_info)

        parts.append(f"Overall expected: {tc.get('expected_result', 'N/A')}")

    text = "\n".join(parts)
    # Truncate to stay within context budget
    return text[:3000]


# ---------------------------------------------------------------------------
# Utility helpers
# ---------------------------------------------------------------------------

def _build_col_map(
    header: List[str], field_aliases: Dict[str, List[str]]
) -> Dict[str, Optional[int]]:
    """Map field names to column indices by matching header against aliases."""
    col_map: Dict[str, Optional[int]] = {}
    for field, aliases in field_aliases.items():
        col_map[field] = None
        for i, h in enumerate(header):
            if any(alias in h for alias in aliases):
                col_map[field] = i
                break
    return col_map


def _cell(row: tuple, col_idx: Optional[int]) -> Any:
    """Safely get a cell value from a row by column index."""
    if col_idx is None or col_idx >= len(row):
        return None
    return row[col_idx]
